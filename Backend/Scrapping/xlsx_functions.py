from openpyxl import Workbook, load_workbook


def save_to_xlsx(data):
    book = Workbook()
    book.remove(book['Sheet'])
    for page, news in sorted(data.items()):
        sheet = book.create_sheet(page)
        for item in range(len(news)):
            sheet[f'A{item + 1}'] = news[item]
    book.save('data.xlsx')


def get_xlsx():
    book = load_workbook('data.xlsx')
    data = {}
    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        column = sheet['A']
        tmp_list = [row.internal_value for row in column]
        data[sheet_name] = tmp_list
    return data
